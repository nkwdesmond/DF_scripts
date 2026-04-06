#!/usr/bin/env python3

import argparse
import os
import re
import sys
import csv
import shlex
from collections import defaultdict

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None


# =========================
# Plugin System
# =========================

PARSERS = {}

def register_parser(name):
    def wrapper(func):
        PARSERS[name] = func
        return func
    return wrapper


# =========================
# Utility Functions
# =========================

def normalize_script(script):
    return os.path.basename(script).lower()


def sanitize_sheet_name(name):
    name = re.sub(r'[\\/*?:\[\]]', "_", name)
    return name[:31] if name else "sheet"


used_sheet_names = set()

def get_unique_sheet_name(name):
    base = sanitize_sheet_name(name)
    new_name = base
    counter = 1

    while new_name in used_sheet_names:
        suffix = f"_{counter}"
        new_name = base[:31 - len(suffix)] + suffix
        counter += 1

    used_sheet_names.add(new_name)
    return new_name


def clean_excel_string(value):
    if not isinstance(value, str):
        return value
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", value)


def is_dirty_line(line):
    return any(ord(c) < 32 and c not in "\t\n\r" for c in line)


def to_hex(s):
    return " ".join(f"{ord(c):02x}" for c in s)


# =========================
# Parsers (Plugins)
# =========================

@register_parser("openfortivpn")
def parse_openfortivpn(line):
    ip = ""
    user = ""
    password = ""

    try:
        tokens = shlex.split(line)
    except Exception:
        tokens = line.split()

    i = 0
    while i < len(tokens):
        t = tokens[i]

        if t in ("-u", "--username"):
            if i + 1 < len(tokens):
                user = tokens[i + 1]
                i += 1
        elif t.startswith("--username="):
            user = t.split("=", 1)[1]

        elif t in ("-p", "--password"):
            if i + 1 < len(tokens):
                password = tokens[i + 1]
                i += 1
        elif t.startswith("--password="):
            password = t.split("=", 1)[1]

        elif not ip:
            m = re.search(r"(?:\d{1,3}\.){3}\d{1,3}", t)
            if m:
                ip = m.group()

        i += 1

    return {
        "Command": line,
        "IP": ip,
        "Username": user,
        "Password": password
    }


# =========================
# Main Logic
# =========================

def main():
    parser = argparse.ArgumentParser(
        description="Parses a text file (e.g. .bash_history) and output a csv. For each unique IP address, this script will provide the count of each type of command ran involving the IP address."
    )

    parser.add_argument("-f", "--file", required=True, help="Input file")
    parser.add_argument("-o", "--output", help="Output file")
    parser.add_argument("--combine", action="store_true",
                        help="Combine the count of all commands for each unique IP address")
    parser.add_argument("-x", "--excel", action="store_true",
                        help="Output Excel workbook with additional individual sheets for each type of command involving an IP address")

    args = parser.parse_args()

    if not os.path.isfile(args.file):
        print("Error: Input file not found.")
        sys.exit(1)

    if args.excel and Workbook is None:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    base = os.path.basename(args.file)
    name, _ = os.path.splitext(base)
    directory = os.path.dirname(args.file)

    if args.output:
        output_file = args.output
    else:
        output_file = (
            os.path.join(directory, f"{name}_processed.xlsx")
            if args.excel else
            os.path.join(directory, f"{name}_ipCmdCount.csv")
        )

    ip_regex = re.compile(r"(?:\d{1,3}\.){3}\d{1,3}")
    py_regex = re.compile(r"[a-zA-Z0-9._-]+\.py")

    if args.combine:
        count = defaultdict(int)
        scripts = defaultdict(set)
    else:
        count = defaultdict(lambda: defaultdict(int))

    script_lines = defaultdict(set)
    misc_lines = set()
    dirty_lines = []

    # =========================
    # Parse Input File
    # =========================

    with open(args.file, "r", errors="ignore") as f:
        for line in f:
            line = line.rstrip()

            # Detect dirty lines
            if is_dirty_line(line):
                cleaned = clean_excel_string(line)

                issues = []
                if "\x00" in line:
                    issues.append("NULL-byte")
                issues.append("control-char")

                dirty_lines.append({
                    "Original": line,
                    "Cleaned": cleaned,
                    "Issue": ",".join(set(issues)),
                    "Hex": to_hex(line)
                })

            ip_match = ip_regex.search(line)

            py_match = py_regex.search(line)
            if py_match:
                script = py_match.group()
            else:
                parts = line.split()
                script = parts[0] if parts else None

            if not script:
                misc_lines.add(line)
                continue

            script = normalize_script(script)

            if ip_match:
                ip = ip_match.group()

                if args.combine:
                    count[ip] += 1
                    scripts[ip].add(script)
                else:
                    count[ip][script] += 1

                script_lines[script].add(line)
            else:
                misc_lines.add(line)

    # =========================
    # CSV Output
    # =========================

    if not args.excel:
        with open(output_file, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["IP", "Count", "Script"])

            if args.combine:
                for ip in count:
                    writer.writerow([ip, count[ip], "\n".join(scripts[ip])])
            else:
                for ip in count:
                    for script in count[ip]:
                        writer.writerow([ip, count[ip][script], script])

        print(f"Report saved to {output_file}")
        return

    # =========================
    # Excel Output
    # =========================

    wb = Workbook()

    # Sheet 1
    ws = wb.active
    ws.title = "ipCmdCount"
    ws.append(["IP", "Count", "Script"])

    if args.combine:
        for ip in count:
            ws.append([
                clean_excel_string(ip),
                count[ip],
                clean_excel_string("\n".join(scripts[ip]))
            ])
    else:
        for ip in count:
            for script in count[ip]:
                ws.append([
                    clean_excel_string(ip),
                    count[ip][script],
                    clean_excel_string(script)
                ])

    # =========================
    # Command Sheets
    # =========================

    for script, lines in sorted(script_lines.items()):
        sheet_name = get_unique_sheet_name(script)
        ws_script = wb.create_sheet(title=sheet_name)

        parser_func = PARSERS.get(script)

        if parser_func:
            parsed_rows = [parser_func(l) for l in sorted(lines)]
            headers = list(parsed_rows[0].keys()) + ["Dirty"]

            ws_script.append(headers)

            for l, row in zip(sorted(lines), parsed_rows):
                ws_script.append([
                    clean_excel_string(row.get(h, "")) for h in headers[:-1]
                ] + [
                    "YES" if is_dirty_line(l) else ""
                ])

        else:
            ws_script.append(["Command", "Dirty"])
            for l in sorted(lines):
                ws_script.append([
                    clean_excel_string(l),
                    "YES" if is_dirty_line(l) else ""
                ])

    # =========================
    # Misc Sheet
    # =========================

    ws_misc = wb.create_sheet(title="miscCommand")
    ws_misc.append(["Entries"])

    for l in sorted(misc_lines):
        ws_misc.append([clean_excel_string(l)])

    # =========================
    # Dirty Lines Sheet
    # =========================

    if dirty_lines:
        ws_dirty = wb.create_sheet(title="dirtyLines")
        ws_dirty.append(["Original", "Cleaned", "Issue", "Hex"])

        for entry in dirty_lines:
            ws_dirty.append([
                clean_excel_string(entry["Original"]),
                clean_excel_string(entry["Cleaned"]),
                entry["Issue"],
                entry["Hex"]
            ])

    wb.save(output_file)

    print(f"Excel report saved to {output_file}")
    print(f"[!] Dirty lines detected: {len(dirty_lines)}")


if __name__ == "__main__":
    main()